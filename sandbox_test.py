

# # Load the workbook
# wb = openpyxl.load_workbook(excel_path, read_only=True)

# # List all sheet names
# sheet_names = wb.sheetnames

# print("Sheet names:")
# for name in sheet_names:
#     print(name)


import openpyxl
excel_path = "/Users/joshualevi/Desktop/ProJets Venture/Base PV Model"


wb = openpyxl.load_workbook(excel_path, read_only=True)
print(wb.sheetnames)